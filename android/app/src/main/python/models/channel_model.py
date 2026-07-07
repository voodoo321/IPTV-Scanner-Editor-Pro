import re
from datetime import datetime
from PySide6 import QtCore, QtGui
from typing import List, Dict, Any
from core.log_manager import global_logger as logger
from ui.styles import AppStyles
from models.channel_mappings import extract_channel_name_from_url

# 自定义数据角色：用于 delegate 从 QModelIndex 取评分（与 ui.quality_bar 中常量保持一致）
QUALITY_SCORE_ROLE = 0x0100 + 100  # Qt.ItemDataRole.UserRole + 100
QUALITY_GRADE_ROLE = 0x0100 + 101  # Qt.ItemDataRole.UserRole + 101


class ChannelListModel(QtCore.QAbstractTableModel):
    """频道列表数据模型"""

    COL_INDEX = 0
    COL_NAME = 1
    COL_RESOLUTION = 2
    COL_URL = 3
    COL_GROUP = 4
    COL_LOGO = 5
    COL_STATUS = 6
    COL_LATENCY = 7
    COL_TVG_ID = 8
    COL_TVG_CHNO = 9
    COL_TVG_SHIFT = 10
    COL_CATCHUP = 11
    COL_CATCHUP_DAYS = 12
    COL_CATCHUP_SOURCE = 13

    COLUMN_FIELD_MAP = {
        COL_NAME: 'name',
        COL_RESOLUTION: 'resolution',
        COL_URL: 'url',
        COL_GROUP: 'group',
        COL_LOGO: 'logo',
        COL_STATUS: 'status',
        COL_LATENCY: 'latency',
        COL_TVG_ID: 'tvg_id',
        COL_TVG_CHNO: 'tvg_chno',
        COL_TVG_SHIFT: 'tvg_shift',
        COL_CATCHUP: 'catchup',
        COL_CATCHUP_DAYS: 'catchup_days',
        COL_CATCHUP_SOURCE: 'catchup_source',
    }

    COLUMN_DEFAULTS = {
            0: None,
            1: '未命名',
            2: '',
            3: '',
            4: '未分类',
            5: '',
            6: '待检测',
            7: '',
            8: '',
            9: '',
            10: '',
            11: '',
            12: '',
            13: '',
        }

    HEADER_TRANSLATION_MAP = {
            "序号": 'serial_number',
            "频道名称": 'channel_name',
            "分辨率": 'resolution',
            "URL": 'channel_url',
            "分组": 'channel_group',
            "Logo地址": 'logo_address',
            "状态": 'status',
            "延迟(ms)": 'latency_ms',
            "TVG-ID": 'tvg_id',
            "TVG频道号": 'tvg_chno',
            "TVG时移": 'tvg_shift',
            "回看": 'catchup',
            "回看天数": 'catchup_days',
            "回看源": 'catchup_source',
        }

    def __init__(self, parent=None):
        super().__init__(parent)
        self.channels: List[Dict[str, Any]] = []
        self.headers = [
            "序号", "频道名称", "分辨率", "URL", "分组",
            "Logo地址", "状态", "延迟(ms)", "TVG-ID",
            "TVG频道号", "TVG时移", "回看", "回看天数", "回看源"
        ]

        # 隐藏列状态：True表示隐藏，False表示显示
        self.hidden_columns = {
            2: False,   # 分辨率
            4: False,   # 分组
            5: False,   # Logo地址
            8: False,   # TVG-ID
            9: False,   # TVG频道号
            10: False,  # TVG时移
            11: False,  # 回看
            12: False,  # 回看天数
            13: False   # 回看源
        }

        # 状态标签更新回调
        self.update_status_label = None

        # 频道名称和分组缓存(用于自动补全)
        self._name_cache = set()
        self._group_cache = set()

        # 语言管理器引用
        self._language_manager = None

        # 原始文件内容：用于保存从文件加载的原始内容
        self._original_file_content = ""

        # 原始文件路径
        self._source_file_path = ""

        # 是否处于隐藏无效项状态
        self._is_hiding_invalid = False

        # 原始频道数据存储（用于非隐藏状态下的保存）
        self._original_channel_data = {}

        # 排序状态
        self._sort_column = -1  # 当前排序列
        self._sort_order = QtCore.Qt.SortOrder.AscendingOrder  # 排序顺序

    def set_language_manager(self, language_manager):
        """设置语言管理器"""
        self._language_manager = language_manager
        # 通知视图更新表头
        self.headerDataChanged.emit(QtCore.Qt.Orientation.Horizontal, 0, len(self.headers) - 1)
        # 强制刷新整个视图以确保表头更新
        self.dataChanged.emit(
            self.index(0, 0),
            self.index(self.rowCount() - 1 if self.rowCount() > 0 else 0, self.columnCount() - 1)
        )

    def rowCount(self, parent=QtCore.QModelIndex()) -> int:
        """返回行数(频道数量)"""
        return len(self.channels)

    def clear(self):
        """清空频道列表"""
        self.beginResetModel()
        self.channels = []
        self.endResetModel()

    def sort_by_indices(self, row_order: list):
        """按指定行顺序重新排列频道"""
        if not row_order or not self.channels:
            return

        max_idx = len(self.channels)
        indexed = {i: self.channels[i] for i in range(max_idx)}
        reordered = []
        for idx in row_order:
            if 0 <= idx < max_idx:
                reordered.append(indexed[idx])

        remaining = [indexed[i] for i in sorted(indexed.keys()) if i not in set(row_order)]
        reordered.extend(remaining)

        self.beginResetModel()
        self.channels = reordered
        self.endResetModel()

    def columnCount(self, parent=QtCore.QModelIndex()) -> int:
        """返回列数（排除隐藏列）"""
        visible_count = len(self.headers)
        for col, hidden in self.hidden_columns.items():
            if hidden and col < len(self.headers):
                visible_count -= 1
        return visible_count

    def _logical_to_actual_column(self, logical_col: int) -> int:
        """将逻辑列索引转换为实际列索引（考虑隐藏列）"""
        if logical_col < 0:
            return -1

        actual_col = 0
        for col in range(len(self.headers)):
            if col in self.hidden_columns and self.hidden_columns[col]:
                continue
            if actual_col == logical_col:
                return col
            actual_col += 1
        return -1

    def _actual_to_logical_column(self, actual_col: int) -> int:
        """将实际列索引转换为逻辑列索引（考虑隐藏列）"""
        if actual_col < 0:
            return -1

        logical_col = 0
        for col in range(len(self.headers)):
            if col == actual_col:
                return logical_col
            if col not in self.hidden_columns or not self.hidden_columns[col]:
                logical_col += 1
        return -1

    def toggle_column_visibility(self, actual_col: int):
        """切换列的隐藏状态"""
        if actual_col in self.hidden_columns:
            self.hidden_columns[actual_col] = not self.hidden_columns[actual_col]
            # 通知视图列数发生变化
            self.beginResetModel()
            self.endResetModel()
            # 通知表头更新
            self.headerDataChanged.emit(QtCore.Qt.Orientation.Horizontal, 0, self.columnCount() - 1)
            return True
        return False

    def is_column_hidden(self, actual_col: int) -> bool:
        """检查列是否隐藏"""
        return self.hidden_columns.get(actual_col, False)

    def data(self, index: QtCore.QModelIndex, role: int = QtCore.Qt.ItemDataRole.DisplayRole):
        """返回单元格数据"""
        if not index.isValid() or not (0 <= index.row() < len(self.channels)):
            return None

        channel = self.channels[index.row()]
        logical_col = index.column()
        actual_col = self._logical_to_actual_column(logical_col)

        if actual_col == -1:
            return None

        if role == QtCore.Qt.ItemDataRole.DisplayRole:
            return self._get_display_data(channel, actual_col, index.row())
        elif role == QtCore.Qt.ItemDataRole.DecorationRole and actual_col == 0:
            return None
        elif role == QtCore.Qt.ItemDataRole.TextAlignmentRole:
            return QtCore.Qt.AlignmentFlag.AlignVCenter | QtCore.Qt.AlignmentFlag.AlignLeft
        elif role == QtCore.Qt.ItemDataRole.BackgroundRole:
            from ui.styles import AppStyles, color_to_qcolor
            colors = AppStyles._get_colors()
            valid = channel.get('valid')
            if valid is False:
                return color_to_qcolor(colors.get('error_background', '#ffdddd'))
            else:
                return color_to_qcolor(colors['table_alternate'])
        elif role == QtCore.Qt.ItemDataRole.ForegroundRole:
            from ui.styles import AppStyles, color_to_qcolor
            colors = AppStyles._get_colors()
            status = channel.get('status', '待检测')
            valid = channel.get('valid')

            if valid is False and status != '待检测':
                return color_to_qcolor(colors.get('error', '#ff6666'))
            elif status == '待检测':
                return color_to_qcolor(colors.get('placeholder', '#999999'))
            else:
                return color_to_qcolor(colors['window_text'])
        elif role == QUALITY_SCORE_ROLE:
            # 显示评分条的条件：
            # - 频道已检测（valid 非 None），按数据计算/读取
            # - 或频道从 M3U 加载且带有 quality_score 字段（持久化的扫描结果）
            valid = channel.get('valid')
            has_persisted_score = 'quality_score' in channel
            if valid is None and not has_persisted_score:
                return None
            score = channel.get('quality_score', None)
            if score is None or score == '':
                if valid is None:
                    return None
                # 评分字段缺失但已检测，按现有数据动态计算（不写入）
                from services.stream_quality_scorer import StreamQualityScorer
                info = StreamQualityScorer.score_from_channel(channel)
                return info.get('total')
            try:
                return float(score)
            except (TypeError, ValueError):
                return None
        elif role == QUALITY_GRADE_ROLE:
            valid = channel.get('valid')
            has_persisted_score = 'quality_score' in channel
            if valid is None and not has_persisted_score:
                return None
            grade = channel.get('quality_grade', None)
            if not grade:
                if valid is None:
                    return None
                from services.stream_quality_scorer import StreamQualityScorer
                info = StreamQualityScorer.score_from_channel(channel)
                return info.get('grade')
            return grade

        return None

    def _get_display_data(self, channel, actual_col, row):
        if actual_col == 0:
            return str(row + 1)
        if actual_col == 5:
            return channel.get('logo', channel.get('logo_url', ''))
        if actual_col == 7:
            latency = channel.get('latency', '')
            if channel.get('valid') is not True:
                return ''
            return str(latency) if latency != '' else ''
        if actual_col == 11:
            catchup_value = channel.get('catchup', '')
            return catchup_value if catchup_value is not None else ''
        field = self.COLUMN_FIELD_MAP.get(actual_col)
        if field:
            default = self.COLUMN_DEFAULTS.get(actual_col, '')
            return channel.get(field, default)
        return None

    def headerData(self, section: int, orientation: QtCore.Qt.Orientation,
                   role: int = QtCore.Qt.ItemDataRole.DisplayRole):
        """返回表头数据"""
        if role != QtCore.Qt.ItemDataRole.DisplayRole:
            return None
        if orientation == QtCore.Qt.Orientation.Horizontal:
            # 将逻辑列转换为实际列
            actual_col = self._logical_to_actual_column(section)
            if actual_col == -1:
                return None

            # 使用语言管理器翻译表头
            header_text = self.headers[actual_col]
            if hasattr(self, '_language_manager') and self._language_manager:
                translation_key = self.HEADER_TRANSLATION_MAP.get(header_text)
                if translation_key:
                    return self._language_manager.tr(translation_key, header_text)
            return header_text
        return str(section + 1) if section > 0 else ""  # 序号列不显示行号

    def flags(self, index: QtCore.QModelIndex) -> QtCore.Qt.ItemFlag:
        """返回项标志"""
        if not index.isValid():
            return QtCore.Qt.ItemFlag.NoItemFlags
        return (QtCore.Qt.ItemFlag.ItemIsEnabled |
                QtCore.Qt.ItemFlag.ItemIsSelectable |
                QtCore.Qt.ItemFlag.ItemIsDragEnabled |
                QtCore.Qt.ItemFlag.ItemIsDropEnabled)

    def supportedDropActions(self) -> QtCore.Qt.DropAction:
        """支持的拖放操作"""
        return QtCore.Qt.DropAction.MoveAction

    def mimeTypes(self) -> List[str]:
        """支持的MIME类型"""
        return ['application/x-channel-row']

    def mimeData(self, indexes: List[QtCore.QModelIndex]) -> QtCore.QMimeData:
        """创建拖放数据"""
        mime_data = QtCore.QMimeData()
        mime_data.setData('application/x-channel-row',
                          str(indexes[0].row()).encode())
        return mime_data

    def dropMimeData(self, data: QtCore.QMimeData, action: QtCore.Qt.DropAction,
                     row: int, column: int, parent: QtCore.QModelIndex) -> bool:
        """处理拖放数据"""
        if not data.hasFormat('application/x-channel-row'):
            return False

        if action == QtCore.Qt.DropAction.IgnoreAction:
            return True

        # 获取拖动源行
        source_row = int(data.data('application/x-channel-row').data().decode())

        # 计算目标行
        if row == -1:
            if parent.isValid():
                row = parent.row()
            else:
                row = self.rowCount()

        # 移动行
        self.moveRow(source_row, row)
        return True

    def moveRow(self, source_row: int, target_row: int) -> bool:
        """移动行到新位置"""
        if source_row == target_row or not (0 <= source_row < len(self.channels)):
            return False

        # 确保目标行在有效范围内
        target_row = min(max(0, target_row), len(self.channels))

        self.beginResetModel()
        channel = self.channels.pop(source_row)
        if target_row > source_row:
            target_row -= 1
        self.channels.insert(target_row, channel)
        self.endResetModel()
        return True

    def get_channel(self, index: int) -> Dict[str, Any]:
        """根据索引获取频道信息"""
        if 0 <= index < len(self.channels):
            return self.channels[index]
        return {}

    def add_channel(self, channel_info: Dict[str, Any], is_from_file: bool = False):
        """添加频道到模型 - 简化版本，只处理单个频道

        Args:
            channel_info: 频道信息字典
            is_from_file: 是否从文件加载的频道（需要保存原始数据）
        """
        # 检查是否已存在相同URL的频道
        existing_index = -1
        for i, channel in enumerate(self.channels):
            if channel.get('url') == channel_info.get('url'):
                existing_index = i
                break

        if existing_index >= 0:
            existing = self.channels[existing_index]
            if existing.get('valid') is not True and channel_info.get('valid') is True:
                self.update_channel(existing_index, channel_info)
            return
        else:
            # 添加新频道
            self.beginInsertRows(QtCore.QModelIndex(), len(self.channels), len(self.channels))
            self.channels.append(channel_info)
            # 更新名称和分组缓存
            if 'name' in channel_info:
                self._name_cache.add(channel_info['name'])
            for g in channel_info.get('_groups', [channel_info.get('group', '')]):
                if g:
                    self._group_cache.add(g)
            self.endInsertRows()

            # 如果是从文件加载的频道，保存原始数据
            if is_from_file:
                url = channel_info.get('url', '')
                if url:
                    # 保存原始数据副本，包含所有M3U属性
                    original_channel = {
                        'name': channel_info.get('name', ''),
                        'group': channel_info.get('group', '未分类'),
                        'tvg_id': channel_info.get('tvg_id', ''),
                        'logo': channel_info.get('logo', ''),
                        'resolution': channel_info.get('resolution', ''),
                        'tvg_chno': channel_info.get('tvg_chno', ''),
                        'tvg_shift': channel_info.get('tvg_shift', ''),
                        'catchup': channel_info.get('catchup', ''),
                        'catchup_days': channel_info.get('catchup_days', ''),
                        'catchup_source': channel_info.get('catchup_source', ''),
                        'url': url
                    }
                    self._original_channel_data[url] = original_channel

    def add_channels(self, channels: List[Dict[str, Any]], is_from_file: bool = False, use_reset: bool = False):
        """批量添加频道到模型

        Args:
            use_reset: 使用 beginResetModel/endResetModel 代替 beginInsertRows/endInsertRows。
                       扫描期间建议 True，避免与 QSortFilterProxyModel 竞态导致 C++ 层崩溃。
        """
        if not channels:
            return
        existing_urls = {ch.get('url') for ch in self.channels}
        new_channels = []
        updates = []
        for ch in channels:
            url = ch.get('url')
            if url not in existing_urls:
                new_channels.append(ch)
                existing_urls.add(url)
            elif ch.get('valid') is True:
                for i, existing in enumerate(self.channels):
                    if existing.get('url') == url and existing.get('valid') is not True:
                        updates.append((i, ch))
                        break
        if not new_channels and not updates:
            return
        if use_reset:
            self.beginResetModel()
            self.channels.extend(new_channels)
            for ch in new_channels:
                if 'name' in ch:
                    self._name_cache.add(ch['name'])
                for g in ch.get('_groups', [ch.get('group', '')]):
                    if g:
                        self._group_cache.add(g)
            for idx, ch in updates:
                self.channels[idx].update(ch)
            self.endResetModel()
        else:
            if new_channels:
                self.beginInsertRows(QtCore.QModelIndex(), len(self.channels), len(self.channels) + len(new_channels) - 1)
                self.channels.extend(new_channels)
                for ch in new_channels:
                    if 'name' in ch:
                        self._name_cache.add(ch['name'])
                    for g in ch.get('_groups', [ch.get('group', '')]):
                        if g:
                            self._group_cache.add(g)
                self.endInsertRows()
            for idx, ch in updates:
                self.update_channel(idx, ch)
        if is_from_file:
            for ch in new_channels:
                url = ch.get('url', '')
                if url:
                    self._original_channel_data[url] = {
                        'name': ch.get('name', ''),
                        'group': ch.get('group', '未分类'),
                        'tvg_id': ch.get('tvg_id', ''),
                        'logo': ch.get('logo', ''),
                        'resolution': ch.get('resolution', ''),
                        'tvg_chno': ch.get('tvg_chno', ''),
                        'tvg_shift': ch.get('tvg_shift', ''),
                        'catchup': ch.get('catchup', ''),
                        'catchup_days': ch.get('catchup_days', ''),
                        'catchup_source': ch.get('catchup_source', ''),
                        'url': url,
                    }

    def hide_invalid(self):
        """隐藏无效频道"""
        if not hasattr(self, '_original_channels'):
            self._original_channels = self.channels.copy()
        self.beginResetModel()
        self.channels = [c for c in self.channels if c.get('valid') is not False]
        self._is_hiding_invalid = True
        self.endResetModel()

    def show_all(self):
        """显示所有频道"""
        self.beginResetModel()
        # 重新加载所有频道数据
        if hasattr(self, '_original_channels'):
            self.channels = self._original_channels.copy()
        self._is_hiding_invalid = False
        self.endResetModel()

    def get_name_suggestions(self) -> List[str]:
        """获取频道名称建议列表"""
        return sorted(self._name_cache)

    def get_group_suggestions(self) -> List[str]:
        """获取分组建议列表"""
        return sorted(self._group_cache)

    def get_all_channel_names(self) -> List[str]:
        """获取所有频道名称列表"""
        return sorted(self._name_cache)

    def _filter_original_content(self) -> str:
        """过滤原始文件内容，只保留有效频道的行"""
        if not self._original_file_content or not self._is_hiding_invalid:
            return ""

        lines = self._original_file_content.splitlines()
        filtered_lines = []
        current_channel_lines = []
        in_channel_block = False
        channel_url = ""

        for line in lines:
            line = line.rstrip()  # 保留行尾空格

            if line.startswith("#EXTINF:"):
                # 开始一个新的频道块
                in_channel_block = True
                current_channel_lines = [line]
                channel_url = ""
            elif line and not line.startswith("#") and in_channel_block:
                # 这是URL行
                channel_url = line
                current_channel_lines.append(line)

                # 检查这个URL对应的频道是否有效
                is_valid = False
                for channel in self.channels:
                    if channel.get('url') == channel_url and channel.get('valid', False):
                        is_valid = True
                        break

                # 如果频道有效，保留这个频道块
                if is_valid:
                    filtered_lines.extend(current_channel_lines)
                    if filtered_lines and filtered_lines[-1] != "":
                        filtered_lines.append("")  # 添加空行分隔

                # 重置状态
                in_channel_block = False
                current_channel_lines = []
                channel_url = ""
            elif in_channel_block:
                # 频道块中的其他行（如注释等）
                current_channel_lines.append(line)
            else:
                # 非频道块的行（如文件头、注释等）
                filtered_lines.append(line)

        # 移除末尾的空行
        while filtered_lines and filtered_lines[-1] == "":
            filtered_lines.pop()

        return "\n".join(filtered_lines)

    def to_m3u(self) -> str:
        """将频道列表转换为M3U格式字符串"""
        from models.channel_mappings import get_channel_info
        from core.log_manager import global_logger as logger
        lines = ["#EXTM3U"]
        for channel in self.channels:
            url = channel.get('url', '')
            if not url:
                continue

            if self._is_hiding_invalid and not channel.get('valid', False):
                continue

            # 检查是否有原始数据可用
            original_channel = self._original_channel_data.get(url)

            if original_channel:
                # 使用原始数据，但保留有效性检测结果
                channel_name = original_channel.get('name', channel.get('name', ''))
                group = original_channel.get('group', channel.get('group', '未分类'))
                tvg_id = original_channel.get('tvg_id', channel.get('tvg_id', ''))
                logo = original_channel.get('logo', channel.get('logo', ''))
                resolution = original_channel.get('resolution', channel.get('resolution', ''))
                # 新增字段
                tvg_chno = original_channel.get('tvg_chno', channel.get('tvg_chno', ''))
                tvg_shift = original_channel.get('tvg_shift', channel.get('tvg_shift', ''))
                catchup = original_channel.get('catchup', channel.get('catchup', ''))
                catchup_days = original_channel.get('catchup_days', channel.get('catchup_days', ''))
                catchup_source = original_channel.get('catchup_source', channel.get('catchup_source', ''))
            else:
                # 使用当前数据（扫描生成的频道）
                channel_name = channel.get('name', '')
                group = channel.get('group', '未分类')
                tvg_id = channel.get('tvg_id', '')
                logo = channel.get('logo', '')
                resolution = channel.get('resolution', '')
                # 新增字段
                tvg_chno = channel.get('tvg_chno', '')
                tvg_shift = channel.get('tvg_shift', '')
                catchup = channel.get('catchup', '')
                catchup_days = channel.get('catchup_days', '')
                catchup_source = channel.get('catchup_source', '')

            # 直接使用频道列表中已有的logo数据，而不是重新调用映射函数
            # 频道列表显示时已经加载了映射后的logo地址，保存在logo或logo_url字段中
            logo_url = channel.get('logo') or channel.get('logo_url') or logo

            # 如果频道列表中没有logo数据，再尝试调用映射函数获取
            if not logo_url:
                channel_info = get_channel_info(channel_name)
                logo_url = channel_info.get('logo_url')
                logger.debug(f"处理频道: {channel_name}, 获取到的信息: {channel_info}")

            # EXTINF行 - 完整格式，包含所有字段
            extinf_parts = ["#EXTINF:-1"]

            # 添加所有属性（排除隐藏列对应的字段）
            if tvg_id and not self.is_column_hidden(8):  # TVG-ID列
                extinf_parts.append(f'tvg-id="{tvg_id}"')
            if channel_name:
                extinf_parts.append(f'tvg-name="{channel_name}"')
            if logo_url and not self.is_column_hidden(5):  # Logo地址列
                extinf_parts.append(f'tvg-logo="{logo_url}"')
            if group and not self.is_column_hidden(4):  # 分组列
                groups = channel.get('_groups', [])
                if groups and len(groups) > 1:
                    group_value = ';'.join(groups)
                else:
                    group_value = group
                extinf_parts.append(f'group-title="{group_value}"')
            if tvg_chno and not self.is_column_hidden(9):  # TVG频道号列
                extinf_parts.append(f'tvg-chno="{tvg_chno}"')
            if tvg_shift and not self.is_column_hidden(10):  # TVG时移列
                extinf_parts.append(f'tvg-shift="{tvg_shift}"')
            if catchup and not self.is_column_hidden(11):  # 回看列
                extinf_parts.append(f'catchup="{catchup}"')
            if catchup_days and not self.is_column_hidden(12):  # 回看天数列
                extinf_parts.append(f'catchup-days="{catchup_days}"')
            if catchup_source and not self.is_column_hidden(13):  # 回看源列
                extinf_parts.append(f'catchup-source="{catchup_source}"')
            catchup_correction = channel.get('catchup_correction', '')
            if catchup_correction:
                extinf_parts.append(f'catchup-correction="{catchup_correction}"')
            if resolution and not self.is_column_hidden(2):  # 分辨率列
                extinf_parts.append(f'resolution="{resolution}"')

            # 流质量评分（持久化到 EXTINF，便于加载后直接显示评分条）
            # valid 为 None（未检测）时不写入
            if channel.get('valid') is not None:
                qs = channel.get('quality_score')
                qg = channel.get('quality_grade')
                # 若评分字段缺失但有 valid，按现有数据计算并回填
                if qs is None:
                    try:
                        from services.stream_quality_scorer import StreamQualityScorer
                        info = StreamQualityScorer.score_from_channel(channel)
                        qs = info.get('total')
                        qg = info.get('grade')
                    except Exception:
                        qs = None
                if qs is not None:
                    extinf_parts.append(f'quality-score="{qs}"')
                if qg:
                    extinf_parts.append(f'quality-grade="{qg}"')

            # 添加频道名称
            extinf_parts.append(f",{channel_name}")

            # 组合成完整的EXTINF行
            extinf = " ".join(extinf_parts)
            lines.append(extinf)

            # URL行
            lines.append(url)

        # 添加来源信息
        from datetime import datetime
        lines.append("\n# Generated by ISEP")
        lines.append(f"# Saved at: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        lines.append("# GitHub: https://github.com/sumingyd/IPTV-Scanner-Editor-Pro")
        return "\n".join(lines)

    def to_txt(self) -> str:
        lines = []
        for channel in self.channels:
            url = channel.get('url', '')
            if not url:
                continue

            original_channel = self._original_channel_data.get(url)

            if original_channel:
                channel_name = original_channel.get('name', channel.get('name', ''))
            else:
                channel_name = channel.get('name', '未命名')

            lines.append(f"{channel_name},{url}")

        content = "\n".join(lines)
        content += f"\n\n# Generated by ISEP"
        content += f"\n# Saved at: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        return content

    def _channels_to_m3u(self, channels: List[Dict[str, Any]]) -> str:
        """将指定频道列表转换为M3U格式字符串"""
        from models.channel_mappings import get_channel_info
        lines = ["#EXTM3U"]
        for channel in channels:
            url = channel.get('url', '')
            if not url:
                continue
            channel_name = channel.get('name', '')
            group = channel.get('group', '未分类')
            tvg_id = channel.get('tvg_id', '')
            logo_url = channel.get('logo') or channel.get('logo_url', '')
            tvg_chno = channel.get('tvg_chno', '')
            tvg_shift = channel.get('tvg_shift', '')
            catchup = channel.get('catchup', '')
            catchup_days = channel.get('catchup_days', '')
            catchup_source = channel.get('catchup_source', '')
            resolution = channel.get('resolution', '')

            if not logo_url:
                channel_info = get_channel_info(channel_name)
                logo_url = channel_info.get('logo_url', '')

            extinf_parts = ["#EXTINF:-1"]
            if tvg_id:
                extinf_parts.append(f'tvg-id="{tvg_id}"')
            if channel_name:
                extinf_parts.append(f'tvg-name="{channel_name}"')
            if logo_url:
                extinf_parts.append(f'tvg-logo="{logo_url}"')
            if group:
                groups = channel.get('_groups', [])
                if groups and len(groups) > 1:
                    group_value = ';'.join(groups)
                else:
                    group_value = group
                extinf_parts.append(f'group-title="{group_value}"')
            if tvg_chno:
                extinf_parts.append(f'tvg-chno="{tvg_chno}"')
            if tvg_shift:
                extinf_parts.append(f'tvg-shift="{tvg_shift}"')
            if catchup:
                extinf_parts.append(f'catchup="{catchup}"')
            if catchup_days:
                extinf_parts.append(f'catchup-days="{catchup_days}"')
            if catchup_source:
                extinf_parts.append(f'catchup-source="{catchup_source}"')
            if resolution:
                extinf_parts.append(f'resolution="{resolution}"')
            extinf_parts.append(f",{channel_name}")
            extinf = " ".join(extinf_parts)
            lines.append(extinf)
            lines.append(url)

        from datetime import datetime
        lines.append("\n# Generated by ISEP")
        lines.append(f"# Saved at: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        lines.append("# GitHub: https://github.com/sumingyd/IPTV-Scanner-Editor-Pro")
        return "\n".join(lines)

    def _channels_to_txt(self, channels: List[Dict[str, Any]]) -> str:
        """将指定频道列表转换为TXT格式字符串"""
        lines = []
        for channel in channels:
            url = channel.get('url', '')
            if not url:
                continue
            channel_name = channel.get('name', '未命名')
            lines.append(f"{channel_name},{url}")
        content = "\n".join(lines)
        content += f"\n\n# Generated by ISEP"
        content += f"\n# Saved at: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        return content

    def to_excel(self, file_path: str) -> bool:
        """将频道列表保存为Excel文件"""
        try:
            from openpyxl import Workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "频道列表"

            # 写入表头
            ws.append(["频道名称", "URL", "分组", "Logo地址", "分辨率", "状态", "延迟(ms)"])

            # 写入数据
            for channel in self.channels:
                ws.append([
                    channel.get('name', '未命名'),
                    channel.get('url', ''),
                    channel.get('group', '未分类'),
                    channel.get('logo', channel.get('logo_url', '')),
                    channel.get('resolution', ''),
                    channel.get('status', '待检测'),
                    channel.get('latency', '')
                ])

            wb.save(file_path)
            return True
        except Exception as e:
            logger.error(f"保存Excel文件失败: {str(e)}", exc_info=True)
            return False

    def from_excel(self, file_path: str) -> bool:
        """从Excel文件加载频道列表"""
        wb = None
        try:
            from openpyxl import load_workbook
            wb = load_workbook(file_path)
            ws = wb.active

            self.beginResetModel()
            self.channels = []
            self._name_cache = set()
            self._group_cache = set()

            # 检查表头是否匹配
            headers = [cell.value for cell in ws[1]]
            expected_headers = ["频道名称", "URL", "分组", "Logo地址", "分辨率", "状态", "延迟(ms)"]
            if headers != expected_headers:
                logger.warning(f"Excel表头不匹配，期望: {expected_headers}，实际: {headers}")

            # 处理数据行
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or not row[0]:  # 跳过空行
                    continue

                try:
                    channel = {
                        'name': str(row[0]) if row[0] else '未命名',
                        'url': str(row[1]) if row[1] else '',
                        'group': str(row[2]) if len(row) > 2 and row[2] else '未分类',
                        'logo': str(row[3]) if len(row) > 3 and row[3] else '',
                        'resolution': str(row[4]) if len(row) > 4 and row[4] else '',
                        'status': str(row[5]) if len(row) > 5 and row[5] else '待检测',
                        'latency': str(row[6]) if len(row) > 6 and row[6] else '',
                        'valid': False
                    }

                    self.channels.append(channel)
                    self._name_cache.add(channel['name'])
                    for g in channel.get('_groups', [channel.get('group', '')]):
                        if g:
                            self._group_cache.add(g)
                    logger.debug(f"成功加载频道: {channel['name']}")
                except Exception as e:
                    logger.error(f"处理Excel行失败: {row}, 错误: {str(e)}")
                    continue

            self.endResetModel()
            logger.info(f"成功从Excel加载 {len(self.channels)} 个频道")
            return True
        except Exception as e:
            logger.error(f"加载Excel文件失败: {str(e)}", exc_info=True)
            return False
        finally:
            if wb:
                wb.close()
                logger.debug("已释放Excel文件资源")

    def removeRow(self, row: int, parent=QtCore.QModelIndex()) -> bool:
        """删除指定行"""
        return self.remove_channel(row, parent)

    def remove_channel(self, row: int, parent=QtCore.QModelIndex()) -> bool:
        """删除指定行的频道"""
        if not (0 <= row < len(self.channels)):
            return False

        channel = self.channels[row]

        self.beginRemoveRows(parent, row, row)

        self.channels.pop(row)

        if 'name' in channel:
            self._name_cache.discard(channel['name'])
        for g in channel.get('_groups', [channel.get('group', '')]):
            if g:
                self._group_cache.discard(g)

        self.endRemoveRows()
        return True

    def set_channel_valid(self, url: str, valid: bool = True) -> bool:
        """设置频道的有效性状态"""
        for i, channel in enumerate(self.channels):
            if channel.get('url') == url:
                self.channels[i]['valid'] = valid
                self.channels[i]['status'] = '有效' if valid else '无效'

                # 通知视图更新特定行
                top_left = self.index(i, 0)
                bottom_right = self.index(i, self.columnCount() - 1)
                self.dataChanged.emit(
                    top_left, bottom_right,
                    [QtCore.Qt.ItemDataRole.DisplayRole,
                     QtCore.Qt.ItemDataRole.BackgroundRole,
                     QtCore.Qt.ItemDataRole.ForegroundRole]
                )
                return True
        return False

    def update_channel_by_url(self, url: str, channel_info: Dict[str, Any]) -> bool:
        """根据URL更新频道信息"""
        for i, channel in enumerate(self.channels):
            if channel.get('url') == url:
                # 记录原始数据用于调试
                old_name = channel.get('name', '')
                old_group = channel.get('group', '')
                old_groups = channel.get('_groups', [])

                # 更新频道数据
                self.channels[i].update(channel_info)

                # 同时更新原始数据存储（如果存在）
                if url in self._original_channel_data:
                    # 更新原始数据存储中的对应频道
                    original_channel = self._original_channel_data[url]
                    # 只更新原始数据中存在的字段
                    for key in ['name', 'group', 'logo', 'tvg_id', 'resolution',
                                'tvg_chno', 'tvg_shift', 'catchup', 'catchup_days', 'catchup_source']:
                        if key in channel_info:
                            original_channel[key] = channel_info[key]

                # 更新名称和分组缓存
                if 'name' in channel_info:
                    self._name_cache.add(channel_info['name'])
                    if old_name and old_name != channel_info['name']:
                        self._name_cache.discard(old_name)
                for g in channel_info.get('_groups', [channel_info.get('group', '')]):
                    if g:
                        self._group_cache.add(g)
                for g in old_groups:
                    if g:
                        self._group_cache.discard(g)

                # 发送数据变化信号，确保特定行更新 - 使用更全面的角色列表
                top_left = self.index(i, 0)
                # 使用实际列数而不是逻辑列数
                bottom_right = self.index(i, len(self.headers) - 1)
                self.dataChanged.emit(top_left, bottom_right, [
                    QtCore.Qt.ItemDataRole.DisplayRole,
                    QtCore.Qt.ItemDataRole.DecorationRole,
                    QtCore.Qt.ItemDataRole.BackgroundRole,
                    QtCore.Qt.ItemDataRole.ForegroundRole
                ])

                # 强制刷新整个视图以确保所有列都更新
                self.layoutChanged.emit()

                return True
        return False

    def update_channel(self, index: int, new_channel: Dict[str, Any]) -> bool:
        """更新指定索引的频道数据"""
        if not (0 <= index < len(self.channels)):
            return False

        # 获取当前频道的URL
        channel = self.channels[index]
        url = channel.get('url', '')
        if not url:
            return False

        # 记录原始数据用于调试
        old_name = channel.get('name', '')
        old_group = channel.get('group', '')
        old_groups = channel.get('_groups', [])

        # 更新频道数据
        self.channels[index].update(new_channel)

        # 同时更新原始数据存储（如果存在）
        if url in self._original_channel_data:
            # 更新原始数据存储中的对应频道
            original_channel = self._original_channel_data[url]
            # 只更新原始数据中存在的字段
            for key in ['name', 'group', 'logo', 'tvg_id', 'resolution',
                        'tvg_chno', 'tvg_shift', 'catchup', 'catchup_days', 'catchup_source']:
                if key in new_channel:
                    original_channel[key] = new_channel[key]

        # 更新名称和分组缓存
        if 'name' in new_channel:
            self._name_cache.add(new_channel['name'])
            if old_name and old_name != new_channel['name']:
                self._name_cache.discard(old_name)
        for g in new_channel.get('_groups', [new_channel.get('group', '')]):
            if g:
                self._group_cache.add(g)
        for g in old_groups:
            if g:
                self._group_cache.discard(g)

        # 发送数据变化信号，确保特定行更新
        top_left = self.index(index, 0)
        bottom_right = self.index(index, self.columnCount() - 1)
        self.dataChanged.emit(
            top_left, bottom_right,
            [QtCore.Qt.ItemDataRole.DisplayRole,
             QtCore.Qt.ItemDataRole.DecorationRole,
             QtCore.Qt.ItemDataRole.BackgroundRole,
             QtCore.Qt.ItemDataRole.ForegroundRole]
            )

        # 强制刷新整个视图以确保所有列都更新
        self.layoutChanged.emit()
        return True

    def update_view(self):
        """批量更新视图"""
        if self.channels:
            top_left = self.index(0, 0)
            bottom_right = self.index(len(self.channels)-1, len(self.headers)-1)
            self.dataChanged.emit(
                top_left, bottom_right,
                [QtCore.Qt.ItemDataRole.DisplayRole,
                 QtCore.Qt.ItemDataRole.DecorationRole,
                 QtCore.Qt.ItemDataRole.BackgroundRole,
                 QtCore.Qt.ItemDataRole.ForegroundRole]
                )
            self.layoutChanged.emit()

    def sort_channels(self, sort_config=None):
        """智能排序频道列表"""
        if not sort_config:
            # 使用默认智能排序
            self._smart_sort()
        else:
            # 使用配置的多条件排序
            self._multi_condition_sort(sort_config)

    def _smart_sort(self):
        """默认智能排序算法"""
        # 定义组名优先级顺序
        group_priority = {
            '央视频道': 0,
            'CETV': 1,
            'CGTN': 2,
            '卫视': 3,
            '国际频道': 4,
            '特色频道': 5,
            '山东频道': 6,
            '市级频道': 7,
            '滨州': 8,
            '德州': 9,
            '东营': 10,
            '菏泽': 11,
            '济南': 12,
            '济宁': 13,
            '聊城': 14,
            '临沂': 15,
            '青岛': 16,
            '日照': 17,
            '泰安': 18,
            '威海': 19,
            '潍坊': 20,
            '烟台': 21,
            '淄博': 22
        }

        # 使用类方法解析分辨率（避免重复定义）

        def get_cctv_number(name):
            """解析CCTV频道编号"""
            if not name:
                return 999

            # 定义精确的频道顺序
            cctv_order = [
                'CCTV-1 综合',
                'CCTV-2 财经',
                'CCTV-3 综艺',
                'CCTV-4 (亚洲)',
                'CCTV-4 (欧洲)',
                'CCTV-4 (美洲)',
                'CCTV-5 体育',
                'CCTV-5+ 体育赛事',
                'CCTV-6 电影',
                'CCTV-7 国防军事',
                'CCTV-8 电视剧',
                'CCTV-9 纪录',
                'CCTV-10 科教',
                'CCTV-11 戏曲',
                'CCTV-12 社会与法',
                'CCTV-13 新闻',
                'CCTV-14 少儿',
                'CCTV-15 音乐',
                'CCTV-16 奥林匹克',
                'CCTV-17 农业农村',
                'CCTV-4K 超高清',
                'CCTV-8K 超高清',
                'CCTV-中视购物',
                '央广购物'
            ]

            # 查找频道在顺序列表中的位置
            for i, channel_name in enumerate(cctv_order):
                if channel_name in name:  # 部分匹配
                    return i

            # 非CCTV频道或未匹配的频道
            return 999

        def get_group_priority(group):
            """获取组名优先级"""
            if not group:
                return len(group_priority) + 1  # 未分类的组放在最后
            for key in group_priority:
                if key in group:  # 部分匹配
                    return group_priority[key]
            return len(group_priority) + 1  # 未匹配的组放在最后

        # 开始排序
        self.beginResetModel()

        # 先按分辨率分组(1920x1080及以上为一组)
        hd_threshold = 1920 * 1080
        self.channels.sort(key=lambda x: (
            self._get_resolution_value(x.get('resolution', '')) < hd_threshold,  # False(高分辨率)在前
            get_group_priority(x.get('group', '')),  # 按组优先级
            get_cctv_number(x.get('name', '')) if '央视频道' in x.get('group', '') else 0,  # CCTV频道特殊排序
            x.get('name', '')  # 按频道名称字母顺序
        ))

        self.endResetModel()

    def _multi_condition_sort(self, sort_config):
        """多条件排序算法"""
        self.beginResetModel()

        # 检查是否是映射文件顺序排序
        if sort_config.get('primary', {}).get('field') == 'mapping_order':
            # 映射文件顺序排序
            self._sort_by_mapping_order()
        else:
            # 构建排序键函数
            def get_sort_key(channel):
                key = []

                # 处理三个优先级
                for priority in ['primary', 'secondary', 'tertiary']:
                    if priority in sort_config:
                        field_config = sort_config[priority]
                        field = field_config['field']
                        method = field_config['method']

                        # 根据字段和方法获取排序值
                        if field == 'group':
                            value = self._get_group_sort_value(channel.get('group', ''),
                                                               method, sort_config.get('group_priority', []))
                        elif field == 'name':
                            value = self._get_name_sort_value(channel.get('name', ''), method)
                        elif field == 'resolution':
                            value = self._get_resolution_sort_value(channel.get('resolution', ''), method)
                        elif field == 'latency':
                            value = self._get_latency_sort_value(channel.get('latency', ''), method)
                        elif field == 'status':
                            value = self._get_status_sort_value(channel.get('status', ''), method)
                        else:
                            value = channel.get(field, '')

                        key.append(value)

                return tuple(key)

            # 执行排序
            self.channels.sort(key=get_sort_key)

        self.endResetModel()

    def _sort_by_mapping_order(self):
        """按映射文件顺序排序"""
        try:
            from models.channel_mappings import mapping_manager

            # 获取映射文件中的映射条目
            mapping_entries = mapping_manager.get_mapping_entries()

            # 如果没有映射条目，直接返回
            if not mapping_entries:
                return

            # 创建映射名称到顺序的映射
            mapping_order = {}
            for i, entry in enumerate(mapping_entries):
                standard_name = entry.get('standard_name', '')
                if standard_name:
                    # 如果同一个标准名称有多个映射条目，使用第一个出现的顺序
                    if standard_name not in mapping_order:
                        mapping_order[standard_name] = i

            # 排序函数
            def get_mapping_order(channel):
                channel_name = channel.get('name', '')
                # 查找映射顺序
                if channel_name in mapping_order:
                    return mapping_order[channel_name]
                # 未找到映射的频道放在最后
                return len(mapping_order) + 1

            # 执行排序
            self.channels.sort(key=get_mapping_order)

        except Exception as e:
            logger.error(f"按映射文件顺序排序失败: {str(e)}", exc_info=True)

    def _get_group_sort_value(self, group, method, group_priority):
        """获取分组的排序值"""
        if method == 'custom':
            # 自定义顺序：使用拖拽排序的优先级
            for i, priority_group in enumerate(group_priority):
                if group == priority_group:
                    return i
            return len(group_priority) + 1  # 不在优先级列表中的放在最后
        elif method == 'alphabetical':
            # 字母顺序
            return group.lower() if group else ''
        elif method == 'reverse_alphabetical':
            # 字母倒序
            return group.lower()[::-1] if group else ''
        else:
            return group.lower() if group else ''

    def _get_name_sort_value(self, name, method):
        """获取名称的排序值"""
        if method == 'alphabetical':
            # 字母顺序
            return name.lower() if name else ''
        elif method == 'reverse_alphabetical':
            # 字母倒序
            return name.lower()[::-1] if name else ''
        elif method == 'pinyin':
            # 拼音顺序（简化实现，实际需要拼音库）
            try:
                # 这里使用简单的拼音首字母排序
                import pypinyin
                pinyin = ''.join([p[0] for p in pypinyin.pinyin(name, style=pypinyin.Style.FIRST_LETTER)])
                return pinyin.lower() if pinyin else name.lower()
            except ImportError as e:
                # 如果没有拼音库，回退到字母顺序
                logger.debug(f"拼音库导入失败，使用字母顺序: {e}")
                return name.lower() if name else ''
            except Exception as e:
                # 其他异常
                logger.warning(f"拼音转换失败，使用字母顺序: {e}")
                return name.lower() if name else ''
        else:
            return name.lower() if name else ''

    def _get_resolution_sort_value(self, resolution, method):
        """获取分辨率的排序值"""
        width, height = self._parse_resolution(resolution)
        total_pixels = width * height

        if method == 'quality_high_to_low':
            # 画质从高到低：像素总数越大，值越小（用于升序排序）
            return -total_pixels
        elif method == 'quality_low_to_high':
            # 画质从低到高：像素总数越小，值越小
            return total_pixels
        elif method == 'width_high_to_low':
            # 宽度从大到小：宽度越大，值越小
            return -width
        elif method == 'width_low_to_high':
            # 宽度从小到大：宽度越小，值越小
            return width
        else:
            return -total_pixels  # 默认画质从高到低

    def _get_latency_sort_value(self, latency, method):
        """获取延迟的排序值"""
        try:
            latency_value = float(latency) if latency else float('inf')
        except (ValueError, TypeError) as e:
            logger.debug(f"解析延迟值失败: {latency}, 错误: {e}")
            latency_value = float('inf')
        except Exception as e:
            logger.warning(f"解析延迟值时发生意外错误: {latency}, 错误: {e}")
            latency_value = float('inf')

        if method == 'low_to_high':
            # 延迟从低到高：延迟越小，值越小
            return latency_value
        elif method == 'high_to_low':
            # 延迟从高到低：延迟越大，值越小
            return -latency_value
        else:
            return latency_value  # 默认延迟从低到高

    def _get_status_sort_value(self, status, method):
        """获取状态的排序值"""
        status_priority = {
            '有效': 0,
            '待检测': 1,
            '无效': 2
        }

        status_value = status_priority.get(status, 3)

        if method == 'valid_first':
            # 有效频道在前：有效频道值更小
            return status_value
        elif method == 'invalid_first':
            # 无效频道在前：无效频道值更小
            return -status_value
        else:
            return status_value  # 默认有效频道在前

    def _parse_resolution(self, resolution):
        """解析分辨率字符串，返回宽度和高度"""
        if not resolution:
            return 0, 0
        try:
            parts = resolution.split('x')
            if len(parts) == 2:
                return int(parts[0]), int(parts[1])
            return 0, 0
        except (ValueError, IndexError) as e:
            logger.debug(f"解析分辨率失败: {resolution}, 错误: {e}")
            return 0, 0
        except Exception as e:
            logger.warning(f"解析分辨率时发生意外错误: {resolution}, 错误: {e}")
            return 0, 0

    def _get_group_priority_value(self, group, group_priority):
        """获取分组的优先级值"""
        if not group:
            return len(group_priority) + 1

        # 在优先级列表中查找分组
        for i, priority_group in enumerate(group_priority):
            if group == priority_group:
                return i

        # 如果分组不在优先级列表中，放在最后
        return len(group_priority) + 1

    def _get_resolution_value(self, resolution):
        """解析分辨率字符串，返回数值用于排序"""
        if not resolution:
            return 0
        try:
            parts = resolution.split('x')
            if len(parts) == 2:
                return int(parts[0]) * int(parts[1])  # 返回像素总数
            return 0
        except (ValueError, IndexError) as e:
            logger.debug(f"解析分辨率失败: {resolution}, 错误: {e}")
            return 0
        except Exception as e:
            logger.warning(f"解析分辨率时发生意外错误: {resolution}, 错误: {e}")
            return 0

    def _get_latency_value(self, latency):
        """解析延迟值，返回数值用于排序"""
        if not latency:
            return float('inf')  # 没有延迟信息的放在最后
        try:
            # 尝试转换为浮点数
            return float(latency)
        except (ValueError, TypeError) as e:
            logger.debug(f"解析延迟值失败: {latency}, 错误: {e}")
            return float('inf')
        except Exception as e:
            logger.warning(f"解析延迟值时发生意外错误: {latency}, 错误: {e}")
            return float('inf')

    def _get_status_value(self, status):
        """获取状态值用于排序"""
        status_priority = {
            '有效': 0,
            '待检测': 1,
            '无效': 2
        }
        return status_priority.get(status, 3)

    def parse_file_content(self, content: str) -> List[Dict[str, Any]]:
        from services.m3u_parser import parse_m3u_content
        channels, header_attrs = parse_m3u_content(content)
        self._last_header_attrs = header_attrs
        self._name_cache = {ch.get('name', '') for ch in channels if ch.get('name')}
        self._group_cache = set()
        for ch in channels:
            for g in ch.get('_groups', [ch.get('group', '')]):
                if g:
                    self._group_cache.add(g)
        return channels

    @staticmethod
    def _is_valid_channel_url(url):
        from services.m3u_parser import is_valid_channel_url
        valid, _ = is_valid_channel_url(url)
        return valid

    def _natural_sort_key(self, s):
        """自然排序键函数，将字符串中的数字部分转换为整数用于排序"""
        if not s:
            return []

        import re
        # 将字符串分割为字母和数字部分
        return [int(text) if text.isdigit() else text.lower()
                for text in re.split(r'(\d+)', str(s))]

    def sort(self, column: int, order: QtCore.Qt.SortOrder = QtCore.Qt.SortOrder.AscendingOrder):
        """按列排序频道列表"""
        if column < 0 or column >= len(self.headers):
            return

        # 保存排序状态
        self._sort_column = column
        self._sort_order = order

        # 开始排序
        self.beginResetModel()

        # 根据列索引选择排序键函数
        if column == 0:  # 序号列 - 按原始顺序
            # 序号列不排序，保持原样
            pass
        elif column == 1:  # 频道名称
            self.channels.sort(key=lambda x: self._natural_sort_key(x.get('name', '')),
                               reverse=(order == QtCore.Qt.SortOrder.DescendingOrder))
        elif column == 2:  # 分辨率
            self.channels.sort(key=lambda x: self._get_resolution_value(x.get('resolution', '')),
                               reverse=(order == QtCore.Qt.SortOrder.DescendingOrder))
        elif column == 3:  # URL
            self.channels.sort(key=lambda x: self._natural_sort_key(x.get('url', '')),
                               reverse=(order == QtCore.Qt.SortOrder.DescendingOrder))
        elif column == 4:  # 分组
            self.channels.sort(key=lambda x: self._natural_sort_key(x.get('group', '')),
                               reverse=(order == QtCore.Qt.SortOrder.DescendingOrder))
        elif column == 5:  # Logo地址
            self.channels.sort(key=lambda x: self._natural_sort_key(x.get('logo', x.get('logo_url', ''))),
                               reverse=(order == QtCore.Qt.SortOrder.DescendingOrder))
        elif column == 6:  # 状态
            self.channels.sort(key=lambda x: self._get_status_value(x.get('status', '')),
                               reverse=(order == QtCore.Qt.SortOrder.DescendingOrder))
        elif column == 7:  # 延迟(ms)
            self.channels.sort(key=lambda x: self._get_latency_value(x.get('latency', '')),
                               reverse=(order == QtCore.Qt.SortOrder.DescendingOrder))
        elif column == 8:  # TVG-ID
            self.channels.sort(key=lambda x: self._natural_sort_key(x.get('tvg_id', '')),
                               reverse=(order == QtCore.Qt.SortOrder.DescendingOrder))
        elif column == 9:  # TVG频道号
            # TVG频道号可能是数字，使用自然排序
            self.channels.sort(key=lambda x: self._natural_sort_key(x.get('tvg_chno', '')),
                               reverse=(order == QtCore.Qt.SortOrder.DescendingOrder))
        elif column == 10:  # TVG时移
            self.channels.sort(key=lambda x: self._natural_sort_key(x.get('tvg_shift', '')),
                               reverse=(order == QtCore.Qt.SortOrder.DescendingOrder))
        elif column == 11:  # 回看
            self.channels.sort(key=lambda x: self._natural_sort_key(x.get('catchup', '')),
                               reverse=(order == QtCore.Qt.SortOrder.DescendingOrder))
        elif column == 12:  # 回看天数
            # 回看天数可能是数字，使用自然排序
            self.channels.sort(key=lambda x: self._natural_sort_key(x.get('catchup_days', '')),
                               reverse=(order == QtCore.Qt.SortOrder.DescendingOrder))
        elif column == 13:  # 回看源
            self.channels.sort(key=lambda x: self._natural_sort_key(x.get('catchup_source', '')),
                               reverse=(order == QtCore.Qt.SortOrder.DescendingOrder))

        self.endResetModel()

    def load_from_file(self, content: str) -> bool:
        """从文件内容加载频道列表"""
        try:
            # 保存原始文件内容
            self._original_file_content = content

            # 使用线程安全的模型重置
            self.beginResetModel()
            self.channels = []
            self._name_cache = set()
            self._group_cache = set()
            # 清空原始数据存储
            self._original_channel_data = {}

            channels = self.parse_file_content(content)
            if channels is None:
                self.endResetModel()
                return False

            for channel in channels:
                existing_index = -1
                for i, ch in enumerate(self.channels):
                    if ch.get('url') == channel.get('url'):
                        existing_index = i
                        break
                if existing_index >= 0:
                    continue
                self.channels.append(channel)
                if 'name' in channel:
                    self._name_cache.add(channel['name'])
                for g in channel.get('_groups', [channel.get('group', '')]):
                    if g:
                        self._group_cache.add(g)
                url = channel.get('url', '')
                if url:
                    original_channel = {
                        'name': channel.get('name', ''),
                        'group': channel.get('group', ''),
                        'tvg_id': channel.get('tvg_id', ''),
                        'logo': channel.get('logo', ''),
                        'resolution': channel.get('resolution', ''),
                        'url': url,
                        '_raw_extinf': channel.get('_raw_extinf', ''),
                        '_all_tags': channel.get('_all_tags', {}),
                    }
                    self._original_channel_data[url] = original_channel

            # 通知UI更新状态标签 - 使用QTimer确保在主线程执行
            if hasattr(self, 'update_status_label') and self.update_status_label:
                from utils.thread_safety import invoke_on_thread
                invoke_on_thread(self, lambda: self.update_status_label("请点击检测有效性按钮"))

            self.endResetModel()

            # 数据加载完成后调整列宽 - 使用QTimer确保在主线程执行
            view = self.parent()
            if view and hasattr(view, 'resizeColumnsToContents'):
                from utils.thread_safety import invoke_on_thread
                invoke_on_thread(self, view.resizeColumnsToContents)

            return True
        except Exception as e:
            logger.error(f"频道模型-加载频道列表失败: {str(e)}", exc_info=True)
            # 确保在异常情况下也调用endResetModel
            try:
                self.endResetModel()
            except RuntimeError as re:
                logger.debug(f"调用endResetModel失败: {re}")
            except Exception as ex:
                logger.warning(f"调用endResetModel时发生意外错误: {ex}")
            return False
